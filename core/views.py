from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth import authenticate, login
from django.db.models import Count, Q
from django.contrib import messages
from django.urls import reverse, reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.http import HttpResponse, JsonResponse
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa
from io import BytesIO
from django.template.loader import get_template
from .models import (
    User, Beneficiary, Case, CaseNote, Assessment, AssessmentQuestion, 
    AssessmentAnswer, Program, BeneficiaryCategory, ReportTemplate, Report,
    Referral, Alert, ActionPlan, BeneficiaryProgress
)
from .serializers import (
    UserSerializer, BeneficiarySerializer, CaseSerializer, CaseNoteSerializer,
    AssessmentSerializer, AssessmentQuestionSerializer, AssessmentAnswerSerializer,
    ProgramSerializer, BeneficiaryCategorySerializer, ReferralSerializer, AlertSerializer,
    ActionPlanSerializer, BeneficiaryProgressSerializer
)

# Authentication Views
def login_view(request):
    """Custom login view that serves the login page at the root URL"""
    if request.user.is_authenticated:
        return redirect('dashboard_redirect')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            next_url = request.GET.get('next', 'dashboard_redirect')
            return redirect(next_url)
        else:
            messages.error(request, "Invalid username or password.")

    return render(request, 'login.html')

# API ViewSets
class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer

class BeneficiaryViewSet(viewsets.ModelViewSet):
    queryset = Beneficiary.objects.all()
    serializer_class = BeneficiarySerializer
    permission_classes = [IsAuthenticated]

class CaseViewSet(viewsets.ModelViewSet):
    queryset = Case.objects.all()
    serializer_class = CaseSerializer
    permission_classes = [IsAuthenticated]

class CaseNoteViewSet(viewsets.ModelViewSet):
    queryset = CaseNote.objects.all()
    serializer_class = CaseNoteSerializer
    permission_classes = [IsAuthenticated]

class AssessmentViewSet(viewsets.ModelViewSet):
    queryset = Assessment.objects.all()
    serializer_class = AssessmentSerializer
    permission_classes = [IsAuthenticated]

class AssessmentQuestionViewSet(viewsets.ModelViewSet):
    queryset = AssessmentQuestion.objects.all()
    serializer_class = AssessmentQuestionSerializer
    permission_classes = [IsAuthenticated]

class AssessmentAnswerViewSet(viewsets.ModelViewSet):
    queryset = AssessmentAnswer.objects.all()
    serializer_class = AssessmentAnswerSerializer
    permission_classes = [IsAuthenticated]
class ProgramViewSet(viewsets.ModelViewSet):
    queryset = Program.objects.all()
    serializer_class = ProgramSerializer
    permission_classes = [IsAuthenticated]

class BeneficiaryCategoryViewSet(viewsets.ModelViewSet):
    queryset = BeneficiaryCategory.objects.all()
    serializer_class = BeneficiaryCategorySerializer
    permission_classes = [IsAuthenticated]

class ReferralViewSet(viewsets.ModelViewSet):
    queryset = Referral.objects.all()
    serializer_class = ReferralSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        # Filter based on user role
        if self.request.user.role in ['field_officer', 'case_manager', 'partner_organisation']:
            queryset = queryset.filter(
                Q(referred_by=self.request.user) | Q(referred_to_user=self.request.user)
            )
        return queryset

    def perform_create(self, serializer):
        serializer.save(referred_by=self.request.user)

class AlertViewSet(viewsets.ModelViewSet):
    queryset = Alert.objects.all()
    serializer_class = AlertSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Show only alerts for the current user
        return Alert.objects.filter(user=self.request.user).order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class ActionPlanViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows action plans to be viewed or edited.
    """
    queryset = ActionPlan.objects.all()
    serializer_class = ActionPlanSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Filter action plans based on user role
        if self.request.user.role == 'case_manager':
            # Case managers can see action plans they created
            return ActionPlan.objects.filter(created_by=self.request.user).order_by('-created_at')
        elif self.request.user.role == 'admin':
            # Admins can see all action plans
            return ActionPlan.objects.all().order_by('-created_at')
        else:
            # Other roles can't see action plans
            return ActionPlan.objects.none()

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class BeneficiaryProgressViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows beneficiary progress to be viewed or edited.
    """
    queryset = BeneficiaryProgress.objects.all()
    serializer_class = BeneficiaryProgressSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Filter progress records based on user role
        if self.request.user.role == 'case_manager':
            # Case managers can see progress for their cases
            case_ids = Case.objects.filter(case_manager=self.request.user).values_list('id', flat=True)
            return BeneficiaryProgress.objects.filter(case__id__in=case_ids).order_by('-date')
        elif self.request.user.role == 'monitoring_and_evaluation':
            # M&E can see all progress records
            return BeneficiaryProgress.objects.all().order_by('-date')
        elif self.request.user.role == 'admin':
            # Admins can see all progress records
            return BeneficiaryProgress.objects.all().order_by('-date')
        else:
            # Other roles can't see progress records
            return BeneficiaryProgress.objects.none()

    def perform_create(self, serializer):
        serializer.save(recorded_by=self.request.user)

# Custom Mixins
class RoleRequiredMixin(UserPassesTestMixin):
    """Mixin that checks if the user has the required role(s)"""
    allowed_roles = []

    def test_func(self):
        return self.request.user.role in self.allowed_roles

class AdminRequiredMixin(RoleRequiredMixin):
    """Mixin that checks if the user is an admin"""
    allowed_roles = ['admin']

class AdminOrCaseManagerRequiredMixin(RoleRequiredMixin):
    """Mixin that checks if the user is an admin or case manager"""
    allowed_roles = ['admin', 'case_manager']

class PartnerOrganisationRequiredMixin(RoleRequiredMixin):
    """Mixin that checks if the user is from a partner organisation"""
    allowed_roles = ['partner_organisation']

class MonitoringAndEvaluationRequiredMixin(RoleRequiredMixin):
    """Mixin that checks if the user is from monitoring and evaluation"""
    allowed_roles = ['monitoring_and_evaluation']

class ProgramDirectorRequiredMixin(RoleRequiredMixin):
    """Mixin that checks if the user is a program director"""
    allowed_roles = ['program_director']

class CanMakeReferralMixin(RoleRequiredMixin):
    """Mixin that checks if the user can make referrals"""
    allowed_roles = ['field_officer', 'case_manager', 'partner_organisation']

class CanReceiveAlertsMixin(RoleRequiredMixin):
    """Mixin that checks if the user can receive alerts"""
    allowed_roles = ['admin', 'program_director']

class CanGenerateReportsMixin(RoleRequiredMixin):
    """Mixin that checks if the user can generate reports"""
    allowed_roles = ['partner_organisation', 'monitoring_and_evaluation', 'program_director']

class CanTrackBeneficiaryProgressMixin(RoleRequiredMixin):
    """Mixin that checks if the user can track beneficiary progress"""
    allowed_roles = ['case_manager', 'monitoring_and_evaluation']

class AnyRoleRequiredMixin(RoleRequiredMixin):
    """Mixin that allows any authenticated user with a valid role"""
    allowed_roles = ['admin', 'case_manager', 'field_officer', 'partner_organisation', 'monitoring_and_evaluation', 'program_director']

# Beneficiary Views
class BeneficiaryListView(LoginRequiredMixin, AnyRoleRequiredMixin, ListView):
    model = Beneficiary
    template_name = 'beneficiaries/beneficiary_list.html'
    context_object_name = 'beneficiaries'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()

        # Apply search filter
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) | 
                Q(address__icontains=search_query)
            )

        # Apply gender filter
        gender = self.request.GET.get('gender', '')
        if gender:
            queryset = queryset.filter(gender=gender)

        # Apply sorting
        sort = self.request.GET.get('sort', 'name')
        queryset = queryset.order_by(sort)

        return queryset

# Case Views
class CaseListView(LoginRequiredMixin, AnyRoleRequiredMixin, ListView):
    model = Case
    template_name = 'cases/case_list.html'
    context_object_name = 'cases'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filter cases based on user role
        if self.request.user.role == 'case_manager':
            queryset = queryset.filter(case_manager=self.request.user)
        elif self.request.user.role == 'field_officer':
            # Field officers can only see cases they've contributed to
            case_notes = CaseNote.objects.filter(created_by=self.request.user).values_list('case_id', flat=True)
            assessments = Assessment.objects.filter(created_by=self.request.user).values_list('case_id', flat=True)
            queryset = queryset.filter(Q(id__in=case_notes) | Q(id__in=assessments)).distinct()

        # Apply search filter
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(title__icontains=search_query) | 
                Q(description__icontains=search_query) |
                Q(beneficiary__name__icontains=search_query)
            )

        # Apply status filter
        status = self.request.GET.get('status', '')
        if status:
            queryset = queryset.filter(status=status)

        # Apply sorting
        sort = self.request.GET.get('sort', '-created_at')
        queryset = queryset.order_by(sort)

        return queryset

class BeneficiaryDetailView(LoginRequiredMixin, AnyRoleRequiredMixin, DetailView):
    model = Beneficiary
    template_name = 'beneficiaries/beneficiary_detail.html'
    context_object_name = 'beneficiary'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        beneficiary = self.get_object()

        # Get related data
        context['cases'] = Case.objects.filter(beneficiary=beneficiary)
        context['active_cases'] = context['cases'].filter(status='open')

        # Get assessments from all cases
        case_ids = context['cases'].values_list('id', flat=True)
        context['assessments'] = Assessment.objects.filter(case__id__in=case_ids)

        # Get case notes from all cases
        context['case_notes'] = CaseNote.objects.filter(case__id__in=case_ids)

        return context

class CaseDetailView(LoginRequiredMixin, AnyRoleRequiredMixin, DetailView):
    model = Case
    template_name = 'cases/case_detail.html'
    context_object_name = 'case'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        case = self.get_object()

        # Get related data
        context['assessments'] = Assessment.objects.filter(case=case).order_by('-created_at')
        context['case_notes'] = CaseNote.objects.filter(case=case).order_by('-created_at')

        return context

class CaseCreateView(LoginRequiredMixin, AdminOrCaseManagerRequiredMixin, CreateView):
    model = Case
    template_name = 'cases/case_form.html'
    fields = ['title', 'beneficiary', 'status', 'description']

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Limit beneficiary choices based on user role
        if self.request.user.role == 'case_manager':
            # Case managers can see all beneficiaries
            pass
        return form

    def form_valid(self, form):
        # Set the case manager to the current user if they are a case manager
        if self.request.user.role == 'case_manager':
            form.instance.case_manager = self.request.user
        messages.success(self.request, f"Case '{form.instance.title}' created successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('case_detail', kwargs={'pk': self.object.pk})

class CaseUpdateView(LoginRequiredMixin, AdminOrCaseManagerRequiredMixin, UpdateView):
    model = Case
    template_name = 'cases/case_form.html'
    fields = ['title', 'beneficiary', 'status', 'description']

    def get_queryset(self):
        queryset = super().get_queryset()
        # Case managers can only update their own cases
        if self.request.user.role == 'case_manager':
            queryset = queryset.filter(case_manager=self.request.user)
        return queryset

    def form_valid(self, form):
        messages.success(self.request, f"Case '{form.instance.title}' updated successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('case_detail', kwargs={'pk': self.object.pk})

class CaseDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = Case
    template_name = 'cases/case_confirm_delete.html'
    context_object_name = 'case'
    success_url = reverse_lazy('case_list')

    def delete(self, request, *args, **kwargs):
        case = self.get_object()
        messages.success(request, f"Case '{case.title}' deleted successfully.")
        return super().delete(request, *args, **kwargs)

# Assessment Views
class AssessmentListView(LoginRequiredMixin, AnyRoleRequiredMixin, ListView):
    model = Assessment
    template_name = 'assessments/assessment_list.html'
    context_object_name = 'assessments'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filter assessments based on user role
        if self.request.user.role == 'case_manager':
            # Case managers can see assessments for their cases
            case_ids = Case.objects.filter(case_manager=self.request.user).values_list('id', flat=True)
            queryset = queryset.filter(case__id__in=case_ids)
        elif self.request.user.role == 'field_officer':
            # Field officers can only see assessments they created
            queryset = queryset.filter(created_by=self.request.user)

        # Apply search filter
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(title__icontains=search_query) | 
                Q(description__icontains=search_query) |
                Q(case__title__icontains=search_query) |
                Q(case__beneficiary__name__icontains=search_query)
            )

        # Apply sorting
        sort = self.request.GET.get('sort', '-created_at')
        queryset = queryset.order_by(sort)

        return queryset

class AssessmentDetailView(LoginRequiredMixin, AnyRoleRequiredMixin, DetailView):
    model = Assessment
    template_name = 'assessments/assessment_detail.html'
    context_object_name = 'assessment'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        assessment = self.get_object()

        # Get related data
        context['questions'] = AssessmentQuestion.objects.filter(assessment=assessment).order_by('order')
        context['answers'] = AssessmentAnswer.objects.filter(question__assessment=assessment)

        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        # Filter based on user role
        if self.request.user.role == 'case_manager':
            # Case managers can see assessments for their cases
            case_ids = Case.objects.filter(case_manager=self.request.user).values_list('id', flat=True)
            queryset = queryset.filter(case__id__in=case_ids)
        elif self.request.user.role == 'field_officer':
            # Field officers can only see assessments they created
            queryset = queryset.filter(created_by=self.request.user)
        return queryset

class AssessmentCreateView(LoginRequiredMixin, AnyRoleRequiredMixin, CreateView):
    model = Assessment
    template_name = 'assessments/assessment_form.html'
    fields = ['title', 'case', 'description', 'amount_received', 'income_amount']

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Limit case choices based on user role
        if self.request.user.role == 'case_manager':
            form.fields['case'].queryset = Case.objects.filter(case_manager=self.request.user)
        elif self.request.user.role == 'field_officer':
            # Field officers can see all cases to create assessments
            # This allows them to contribute to cases even if they haven't before
            form.fields['case'].queryset = Case.objects.all()
        return form

    def form_valid(self, form):
        # Set the creator to the current user
        form.instance.created_by = self.request.user

        # Save the form to get the assessment instance
        response = super().form_valid(form)

        beneficiary = self.object.case.beneficiary
        updated = False

        # Check if the beneficiary can be promoted to another category
        can_promote_category, new_category = self.object.check_category_promotion()

        if can_promote_category and new_category:
            old_category = beneficiary.category
            beneficiary.category = new_category
            updated = True

            # Add a success message about the category promotion
            if old_category:
                messages.success(self.request, 
                    f"Beneficiary '{beneficiary.name}' has been promoted from '{old_category.name}' to '{new_category.name}' category based on income assessment.")
            else:
                messages.success(self.request, 
                    f"Beneficiary '{beneficiary.name}' has been assigned to '{new_category.name}' category based on income assessment.")

        # Check if the beneficiary can be promoted to another program
        can_promote_program, new_program = self.object.check_program_promotion()

        if can_promote_program and new_program:
            old_program = beneficiary.program
            beneficiary.program = new_program
            updated = True

            # Add a success message about the program promotion
            messages.success(self.request, 
                f"Beneficiary '{beneficiary.name}' has been promoted from '{old_program.name}' to '{new_program.name}' program.")

        # Save the beneficiary if it was updated
        if updated:
            beneficiary.save()

        messages.success(self.request, f"Assessment '{form.instance.title}' created successfully.")
        return response

    def get_success_url(self):
        return reverse('assessment_detail', kwargs={'pk': self.object.pk})

class AssessmentUpdateView(LoginRequiredMixin, AdminOrCaseManagerRequiredMixin, UpdateView):
    model = Assessment
    template_name = 'assessments/assessment_form.html'
    fields = ['title', 'case', 'description', 'amount_received', 'income_amount']

    def get_queryset(self):
        queryset = super().get_queryset()
        # Case managers can only update assessments for their cases
        if self.request.user.role == 'case_manager':
            case_ids = Case.objects.filter(case_manager=self.request.user).values_list('id', flat=True)
            queryset = queryset.filter(case__id__in=case_ids)
        return queryset

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Limit case choices based on user role
        if self.request.user.role == 'case_manager':
            form.fields['case'].queryset = Case.objects.filter(case_manager=self.request.user)
        return form

    def form_valid(self, form):
        # Save the form to get the assessment instance
        response = super().form_valid(form)

        beneficiary = self.object.case.beneficiary
        updated = False

        # Check if the beneficiary can be promoted to another category
        can_promote_category, new_category = self.object.check_category_promotion()

        if can_promote_category and new_category:
            old_category = beneficiary.category
            beneficiary.category = new_category
            updated = True

            # Add a success message about the category promotion
            if old_category:
                messages.success(self.request, 
                    f"Beneficiary '{beneficiary.name}' has been promoted from '{old_category.name}' to '{new_category.name}' category based on income assessment.")
            else:
                messages.success(self.request, 
                    f"Beneficiary '{beneficiary.name}' has been assigned to '{new_category.name}' category based on income assessment.")

        # Check if the beneficiary can be promoted to another program
        can_promote_program, new_program = self.object.check_program_promotion()

        if can_promote_program and new_program:
            old_program = beneficiary.program
            beneficiary.program = new_program
            updated = True

            # Add a success message about the program promotion
            messages.success(self.request, 
                f"Beneficiary '{beneficiary.name}' has been promoted from '{old_program.name}' to '{new_program.name}' program.")

        # Save the beneficiary if it was updated
        if updated:
            beneficiary.save()

        messages.success(self.request, f"Assessment '{form.instance.title}' updated successfully.")
        return response

    def get_success_url(self):
        return reverse('assessment_detail', kwargs={'pk': self.object.pk})

class AssessmentDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = Assessment
    template_name = 'assessments/assessment_confirm_delete.html'
    context_object_name = 'assessment'
    success_url = reverse_lazy('assessment_list')

    def delete(self, request, *args, **kwargs):
        assessment = self.get_object()
        messages.success(request, f"Assessment '{assessment.title}' deleted successfully.")
        return super().delete(request, *args, **kwargs)

# Case Note Views
class CaseNoteListView(LoginRequiredMixin, AnyRoleRequiredMixin, ListView):
    model = CaseNote
    template_name = 'case_notes/case_note_list.html'
    context_object_name = 'case_notes'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filter case notes based on user role
        if self.request.user.role == 'case_manager':
            # Case managers can see notes for their cases
            case_ids = Case.objects.filter(case_manager=self.request.user).values_list('id', flat=True)
            queryset = queryset.filter(case__id__in=case_ids)
        elif self.request.user.role == 'field_officer':
            # Field officers can only see notes they created
            queryset = queryset.filter(created_by=self.request.user)

        # Apply search filter
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(content__icontains=search_query) | 
                Q(case__title__icontains=search_query) |
                Q(case__beneficiary__name__icontains=search_query)
            )

        # Apply sorting
        sort = self.request.GET.get('sort', '-created_at')
        queryset = queryset.order_by(sort)

        return queryset

class CaseNoteDetailView(LoginRequiredMixin, AnyRoleRequiredMixin, DetailView):
    model = CaseNote
    template_name = 'case_notes/case_note_detail.html'
    context_object_name = 'case_note'

    def get_queryset(self):
        queryset = super().get_queryset()
        # Filter based on user role
        if self.request.user.role == 'case_manager':
            # Case managers can see notes for their cases
            case_ids = Case.objects.filter(case_manager=self.request.user).values_list('id', flat=True)
            queryset = queryset.filter(case__id__in=case_ids)
        elif self.request.user.role == 'field_officer':
            # Field officers can only see notes they created
            queryset = queryset.filter(created_by=self.request.user)
        return queryset

class CaseNoteCreateView(LoginRequiredMixin, AnyRoleRequiredMixin, CreateView):
    model = CaseNote
    template_name = 'case_notes/case_note_form.html'
    fields = ['case', 'content']

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Limit case choices based on user role
        if self.request.user.role == 'case_manager':
            form.fields['case'].queryset = Case.objects.filter(case_manager=self.request.user)
        elif self.request.user.role == 'field_officer':
            # Field officers can see all cases to create notes
            # This allows them to contribute to cases even if they haven't before
            form.fields['case'].queryset = Case.objects.all()
        return form

    def form_valid(self, form):
        # Set the creator to the current user
        form.instance.created_by = self.request.user
        messages.success(self.request, "Case note created successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('case_note_detail', kwargs={'pk': self.object.pk})

class CaseNoteUpdateView(LoginRequiredMixin, AnyRoleRequiredMixin, UpdateView):
    model = CaseNote
    template_name = 'case_notes/case_note_form.html'
    fields = ['case', 'content']

    def get_queryset(self):
        queryset = super().get_queryset()
        # Users can only update their own notes
        queryset = queryset.filter(created_by=self.request.user)
        return queryset

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Limit case choices based on user role
        if self.request.user.role == 'case_manager':
            form.fields['case'].queryset = Case.objects.filter(case_manager=self.request.user)
        elif self.request.user.role == 'field_officer':
            # Field officers can see all cases when updating notes
            # This allows them to move their notes to different cases if needed
            form.fields['case'].queryset = Case.objects.all()
        return form

    def form_valid(self, form):
        messages.success(self.request, "Case note updated successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('case_note_detail', kwargs={'pk': self.object.pk})

class CaseNoteDeleteView(LoginRequiredMixin, AnyRoleRequiredMixin, DeleteView):
    model = CaseNote
    template_name = 'case_notes/case_note_confirm_delete.html'
    context_object_name = 'case_note'
    success_url = reverse_lazy('case_note_list')

    def get_queryset(self):
        queryset = super().get_queryset()
        # Users can only delete their own notes, admins can delete any
        if self.request.user.role != 'admin':
            queryset = queryset.filter(created_by=self.request.user)
        return queryset

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Case note deleted successfully.")
        return super().delete(request, *args, **kwargs)

# Visit Views (for Field Officers)
class VisitListView(CaseNoteListView):
    """Alias for CaseNoteListView with a different template for Field Officers"""
    template_name = 'visits/visit_list.html'

    def get_queryset(self):
        # Ensure only field officers can access this view
        if self.request.user.role != 'field_officer':
            return CaseNote.objects.none()
        return super().get_queryset()

class BeneficiaryCreateView(LoginRequiredMixin, AdminOrCaseManagerRequiredMixin, CreateView):
    model = Beneficiary
    template_name = 'beneficiaries/beneficiary_form.html'
    fields = ['name', 'gender', 'dob', 'address', 'category', 'program']

    def get_success_url(self):
        return reverse('beneficiary_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, f"Beneficiary '{form.instance.name}' created successfully.")
        return super().form_valid(form)

class BeneficiaryUpdateView(LoginRequiredMixin, AdminOrCaseManagerRequiredMixin, UpdateView):
    model = Beneficiary
    template_name = 'beneficiaries/beneficiary_form.html'
    fields = ['name', 'gender', 'dob', 'address', 'category', 'program']

    def get_success_url(self):
        return reverse('beneficiary_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, f"Beneficiary '{form.instance.name}' updated successfully.")
        return super().form_valid(form)

class BeneficiaryDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = Beneficiary
    template_name = 'beneficiaries/beneficiary_confirm_delete.html'
    context_object_name = 'beneficiary'
    success_url = reverse_lazy('beneficiary_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        beneficiary = self.get_object()

        # Get related data
        context['cases'] = Case.objects.filter(beneficiary=beneficiary)

        # Get assessments from all cases
        case_ids = context['cases'].values_list('id', flat=True)
        context['assessments'] = Assessment.objects.filter(case__id__in=case_ids)

        # Get case notes from all cases
        context['case_notes'] = CaseNote.objects.filter(case__id__in=case_ids)

        return context

    def delete(self, request, *args, **kwargs):
        beneficiary = self.get_object()
        messages.success(request, f"Beneficiary '{beneficiary.name}' deleted successfully.")
        return super().delete(request, *args, **kwargs)

# Dashboard Views
@login_required
def dashboard_redirect(request):
    """Redirects users to the appropriate dashboard based on their role"""
    if request.user.role == 'admin':
        return redirect('admin_dashboard')
    elif request.user.role == 'case_manager':
        return redirect('case_manager_dashboard')
    elif request.user.role == 'field_officer':
        return redirect('field_officer_dashboard')
    elif request.user.role == 'partner_organisation':
        return redirect('partner_organisation_dashboard')
    elif request.user.role == 'monitoring_and_evaluation':
        return redirect('monitoring_and_evaluation_dashboard')
    elif request.user.role == 'program_director':
        return redirect('program_director_dashboard')
    else:
        messages.error(request, "Your account doesn't have a valid role assigned.")
        return redirect('login')

@login_required
def admin_dashboard(request):
    """Admin dashboard view"""
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access the admin dashboard.")
        return redirect('dashboard_redirect')

    # Get counts for dashboard cards
    user_count = User.objects.count()
    beneficiary_count = Beneficiary.objects.count()
    active_case_count = Case.objects.filter(status='open').count()
    assessment_count = Assessment.objects.count()
    program_count = Program.objects.count()
    category_count = BeneficiaryCategory.objects.count()

    # Get recent cases
    recent_cases = Case.objects.all().order_by('-created_at')[:10]

    # Get recent programs and categories
    recent_programs = Program.objects.all().order_by('-id')[:5]
    recent_categories = BeneficiaryCategory.objects.all().order_by('-id')[:5]

    context = {
        'user_count': user_count,
        'beneficiary_count': beneficiary_count,
        'active_case_count': active_case_count,
        'assessment_count': assessment_count,
        'program_count': program_count,
        'category_count': category_count,
        'recent_cases': recent_cases,
        'recent_programs': recent_programs,
        'recent_categories': recent_categories,
    }

    return render(request, 'dashboard/admin_dashboard.html', context)

@login_required
def case_manager_dashboard(request):
    """Case manager dashboard view"""
    if request.user.role != 'case_manager':
        messages.error(request, "You don't have permission to access the case manager dashboard.")
        return redirect('dashboard_redirect')

    # Get counts for dashboard cards
    active_case_count = Case.objects.filter(case_manager=request.user, status='open').count()
    beneficiary_count = Beneficiary.objects.count()
    assessment_count = Assessment.objects.filter(created_by=request.user).count()
    program_count = Program.objects.count()
    category_count = BeneficiaryCategory.objects.count()
    action_plan_count = ActionPlan.objects.filter(created_by=request.user).count()

    # Get case IDs for this case manager
    case_ids = Case.objects.filter(case_manager=request.user).values_list('id', flat=True)

    # Get progress records count for this case manager's cases
    progress_count = BeneficiaryProgress.objects.filter(case__id__in=case_ids).count()

    # Get cases assigned to this case manager
    my_cases = Case.objects.filter(case_manager=request.user).order_by('-created_at')[:10]

    # Get recent case notes by this user
    recent_notes = CaseNote.objects.filter(created_by=request.user).order_by('-created_at')[:10]

    # Get recent action plans by this user
    recent_action_plans = ActionPlan.objects.filter(created_by=request.user).order_by('-created_at')[:10]

    # Get recent progress records for this case manager's cases
    recent_progress = BeneficiaryProgress.objects.filter(case__id__in=case_ids).order_by('-date')[:10]

    # Get recent programs and categories
    recent_programs = Program.objects.all().order_by('-id')[:5]
    recent_categories = BeneficiaryCategory.objects.all().order_by('-id')[:5]

    context = {
        'active_case_count': active_case_count,
        'beneficiary_count': beneficiary_count,
        'assessment_count': assessment_count,
        'program_count': program_count,
        'category_count': category_count,
        'action_plan_count': action_plan_count,
        'progress_count': progress_count,
        'my_cases': my_cases,
        'recent_notes': recent_notes,
        'recent_action_plans': recent_action_plans,
        'recent_progress': recent_progress,
        'recent_programs': recent_programs,
        'recent_categories': recent_categories,
    }

    return render(request, 'dashboard/case_manager_dashboard.html', context)

@login_required
def field_officer_dashboard(request):
    """Field officer dashboard view"""
    if request.user.role != 'field_officer':
        messages.error(request, "You don't have permission to access the field officer dashboard.")
        return redirect('dashboard_redirect')

    # Get counts for dashboard cards
    assessment_count = Assessment.objects.filter(created_by=request.user).count()
    case_note_count = CaseNote.objects.filter(created_by=request.user).count()
    beneficiary_count = Beneficiary.objects.count()
    referral_count = Referral.objects.filter(referred_by=request.user).count()

    # Get recent assessments by this user
    recent_assessments = Assessment.objects.filter(created_by=request.user).order_by('-created_at')[:10]

    # Get recent case notes by this user
    recent_notes = CaseNote.objects.filter(created_by=request.user).order_by('-created_at')[:10]

    # Get recent referrals by this user
    recent_referrals = Referral.objects.filter(referred_by=request.user).order_by('-created_at')[:10]

    context = {
        'assessment_count': assessment_count,
        'case_note_count': case_note_count,
        'beneficiary_count': beneficiary_count,
        'referral_count': referral_count,
        'recent_assessments': recent_assessments,
        'recent_notes': recent_notes,
        'recent_referrals': recent_referrals,
    }

    return render(request, 'dashboard/field_officer_dashboard.html', context)


@login_required
def partner_organisation_dashboard(request):
    """Partner organisation dashboard view"""
    if request.user.role != 'partner_organisation':
        messages.error(request, "You don't have permission to access the partner organisation dashboard.")
        return redirect('dashboard_redirect')

    # Get counts for dashboard cards
    referral_count = Referral.objects.filter(referred_by=request.user).count()
    received_referral_count = Referral.objects.filter(referred_to_user=request.user).count()
    report_count = Report.objects.filter(created_by=request.user).count()
    beneficiary_count = Beneficiary.objects.count()

    # Get recent referrals by this user
    recent_referrals = Referral.objects.filter(referred_by=request.user).order_by('-created_at')[:10]

    # Get recent referrals to this user
    recent_received_referrals = Referral.objects.filter(referred_to_user=request.user).order_by('-created_at')[:10]

    # Get recent reports by this user
    recent_reports = Report.objects.filter(created_by=request.user).order_by('-created_at')[:10]

    context = {
        'referral_count': referral_count,
        'received_referral_count': received_referral_count,
        'report_count': report_count,
        'beneficiary_count': beneficiary_count,
        'recent_referrals': recent_referrals,
        'recent_received_referrals': recent_received_referrals,
        'recent_reports': recent_reports,
    }

    return render(request, 'dashboard/partner_organisation_dashboard.html', context)


@login_required
def monitoring_and_evaluation_dashboard(request):
    """Monitoring and evaluation dashboard view"""
    if request.user.role != 'monitoring_and_evaluation':
        messages.error(request, "You don't have permission to access the monitoring and evaluation dashboard.")
        return redirect('dashboard_redirect')

    # Get counts for dashboard cards
    beneficiary_count = Beneficiary.objects.count()
    case_count = Case.objects.count()
    assessment_count = Assessment.objects.count()
    report_count = Report.objects.filter(created_by=request.user).count()
    action_plan_count = ActionPlan.objects.count()
    progress_count = BeneficiaryProgress.objects.count()

    # Get recent cases
    recent_cases = Case.objects.all().order_by('-created_at')[:10]

    # Get recent assessments
    recent_assessments = Assessment.objects.all().order_by('-created_at')[:10]

    # Get recent reports by this user
    recent_reports = Report.objects.filter(created_by=request.user).order_by('-created_at')[:10]

    # Get recent action plans
    recent_action_plans = ActionPlan.objects.all().order_by('-created_at')[:10]

    # Get recent progress records
    recent_progress = BeneficiaryProgress.objects.all().order_by('-date')[:10]

    context = {
        'beneficiary_count': beneficiary_count,
        'case_count': case_count,
        'assessment_count': assessment_count,
        'report_count': report_count,
        'action_plan_count': action_plan_count,
        'progress_count': progress_count,
        'recent_cases': recent_cases,
        'recent_assessments': recent_assessments,
        'recent_reports': recent_reports,
        'recent_action_plans': recent_action_plans,
        'recent_progress': recent_progress,
    }

    return render(request, 'dashboard/monitoring_and_evaluation_dashboard.html', context)


@login_required
def program_director_dashboard(request):
    """Program director dashboard view"""
    if request.user.role != 'program_director':
        messages.error(request, "You don't have permission to access the program director dashboard.")
        return redirect('dashboard_redirect')

    # Get counts for dashboard cards
    beneficiary_count = Beneficiary.objects.count()
    program_count = Program.objects.count()
    case_count = Case.objects.count()
    report_count = Report.objects.filter(created_by=request.user).count()
    alert_count = Alert.objects.filter(user=request.user, is_read=False).count()

    # Get recent programs
    recent_programs = Program.objects.all().order_by('-id')[:10]

    # Get recent reports by this user
    recent_reports = Report.objects.filter(created_by=request.user).order_by('-created_at')[:10]

    # Get unread alerts for this user
    unread_alerts = Alert.objects.filter(user=request.user, is_read=False).order_by('-created_at')[:10]

    context = {
        'beneficiary_count': beneficiary_count,
        'program_count': program_count,
        'case_count': case_count,
        'report_count': report_count,
        'alert_count': alert_count,
        'recent_programs': recent_programs,
        'recent_reports': recent_reports,
        'unread_alerts': unread_alerts,
    }

    return render(request, 'dashboard/program_director_dashboard.html', context)

# Referral Views
class ReferralListView(LoginRequiredMixin, AnyRoleRequiredMixin, ListView):
    model = Referral
    template_name = 'referrals/referral_list.html'
    context_object_name = 'referrals'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filter referrals based on user role
        if self.request.user.role in ['field_officer', 'case_manager', 'partner_organisation']:
            # These roles can see referrals they made
            queryset = queryset.filter(
                Q(referred_by=self.request.user) | Q(referred_to_user=self.request.user)
            )
        elif self.request.user.role == 'monitoring_and_evaluation':
            # M&E can see all referrals for tracking
            pass
        elif self.request.user.role == 'program_director':
            # Program directors can see all referrals
            pass

        # Apply search filter
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(beneficiary__name__icontains=search_query) | 
                Q(referred_to_organization__icontains=search_query) |
                Q(reason__icontains=search_query)
            )

        # Apply status filter
        status = self.request.GET.get('status', '')
        if status:
            queryset = queryset.filter(status=status)

        # Apply sorting
        sort = self.request.GET.get('sort', '-created_at')
        queryset = queryset.order_by(sort)

        return queryset

class ReferralDetailView(LoginRequiredMixin, AnyRoleRequiredMixin, DetailView):
    model = Referral
    template_name = 'referrals/referral_detail.html'
    context_object_name = 'referral'

    def get_queryset(self):
        queryset = super().get_queryset()
        # Filter based on user role
        if self.request.user.role in ['field_officer', 'case_manager', 'partner_organisation']:
            queryset = queryset.filter(
                Q(referred_by=self.request.user) | Q(referred_to_user=self.request.user)
            )
        return queryset

class ReferralCreateView(LoginRequiredMixin, CanMakeReferralMixin, CreateView):
    model = Referral
    template_name = 'referrals/referral_form.html'
    fields = ['beneficiary', 'case', 'referred_to_organization', 'referred_to_user', 'reason', 'notes']

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Limit case choices based on user role
        if self.request.user.role == 'case_manager':
            form.fields['case'].queryset = Case.objects.filter(case_manager=self.request.user)
        elif self.request.user.role == 'field_officer':
            # Field officers can see all cases to create referrals
            form.fields['case'].queryset = Case.objects.all()

        # Limit referred_to_user choices to partner organisations and case managers
        form.fields['referred_to_user'].queryset = User.objects.filter(
            role__in=['partner_organisation', 'case_manager']
        )
        return form

    def form_valid(self, form):
        # Set the referrer to the current user
        form.instance.referred_by = self.request.user
        messages.success(self.request, f"Referral for {form.instance.beneficiary.name} created successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('referral_detail', kwargs={'pk': self.object.pk})

class ReferralUpdateView(LoginRequiredMixin, AnyRoleRequiredMixin, UpdateView):
    model = Referral
    template_name = 'referrals/referral_form.html'
    fields = ['status', 'notes']

    def get_queryset(self):
        queryset = super().get_queryset()
        # Users can only update referrals they made or received
        queryset = queryset.filter(
            Q(referred_by=self.request.user) | Q(referred_to_user=self.request.user)
        )
        return queryset

    def form_valid(self, form):
        messages.success(self.request, f"Referral for {form.instance.beneficiary.name} updated successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('referral_detail', kwargs={'pk': self.object.pk})

class ReferralDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = Referral
    template_name = 'referrals/referral_confirm_delete.html'
    context_object_name = 'referral'
    success_url = reverse_lazy('referral_list')

    def delete(self, request, *args, **kwargs):
        referral = self.get_object()
        messages.success(request, f"Referral for {referral.beneficiary.name} deleted successfully.")
        return super().delete(request, *args, **kwargs)

# Alert Views
class AlertListView(LoginRequiredMixin, CanReceiveAlertsMixin, ListView):
    model = Alert
    template_name = 'alerts/alert_list.html'
    context_object_name = 'alerts'
    paginate_by = 10

    def get_queryset(self):
        # Show only alerts for the current user
        return Alert.objects.filter(user=self.request.user).order_by('-created_at')

class AlertDetailView(LoginRequiredMixin, CanReceiveAlertsMixin, DetailView):
    model = Alert
    template_name = 'alerts/alert_detail.html'
    context_object_name = 'alert'

    def get_queryset(self):
        # Show only alerts for the current user
        return Alert.objects.filter(user=self.request.user)

    def get(self, request, *args, **kwargs):
        # Mark the alert as read when viewed
        response = super().get(request, *args, **kwargs)
        alert = self.object
        if not alert.is_read:
            alert.is_read = True
            alert.save()
        return response

class AlertDeleteView(LoginRequiredMixin, CanReceiveAlertsMixin, DeleteView):
    model = Alert
    template_name = 'alerts/alert_confirm_delete.html'
    context_object_name = 'alert'
    success_url = reverse_lazy('alert_list')

    def get_queryset(self):
        # Users can only delete their own alerts
        return Alert.objects.filter(user=self.request.user)

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Alert deleted successfully.")
        return super().delete(request, *args, **kwargs)

@login_required
def mark_all_alerts_read(request):
    """Mark all alerts for the current user as read"""
    if request.method == 'POST':
        Alert.objects.filter(user=request.user, is_read=False).update(is_read=True)
        messages.success(request, "All alerts marked as read.")
    return redirect('alert_list')

# Program Views
class ProgramListView(LoginRequiredMixin, AnyRoleRequiredMixin, ListView):
    model = Program
    template_name = 'programs/program_list.html'
    context_object_name = 'programs'
    paginate_by = 10

class ProgramDetailView(LoginRequiredMixin, AnyRoleRequiredMixin, DetailView):
    model = Program
    template_name = 'programs/program_detail.html'
    context_object_name = 'program'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        program = self.get_object()

        # Calculate total monthly disbursement
        beneficiary_count = program.beneficiaries.count()
        total_monthly_disbursement = program.monthly_amount * beneficiary_count
        context['total_monthly_disbursement'] = total_monthly_disbursement

        # Get related programs (programs that beneficiaries in this program might be eligible for)
        related_programs = Program.objects.exclude(id=program.id)
        if program.next_program:
            # Ensure the next program is first in the list
            related_programs = list(related_programs)
            if program.next_program in related_programs:
                related_programs.remove(program.next_program)
            related_programs.insert(0, program.next_program)
        context['related_programs'] = related_programs[:5]  # Limit to 5 related programs

        return context

class ProgramCreateView(LoginRequiredMixin, AdminOrCaseManagerRequiredMixin, CreateView):
    model = Program
    fields = ['name', 'description', 'monthly_amount', 'next_program']
    template_name = 'programs/program_form.html'
    success_url = '/programs/'

    def form_valid(self, form):
        messages.success(self.request, f"Program '{form.instance.name}' created successfully.")
        return super().form_valid(form)

class ProgramUpdateView(LoginRequiredMixin, AdminOrCaseManagerRequiredMixin, UpdateView):
    model = Program
    fields = ['name', 'description', 'monthly_amount', 'next_program']
    template_name = 'programs/program_form.html'
    success_url = '/programs/'

    def form_valid(self, form):
        messages.success(self.request, f"Program '{form.instance.name}' updated successfully.")
        return super().form_valid(form)

class ProgramDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = Program
    template_name = 'programs/program_confirm_delete.html'
    success_url = '/programs/'

    def delete(self, request, *args, **kwargs):
        program = self.get_object()
        messages.success(request, f"Program '{program.name}' deleted successfully.")
        return super().delete(request, *args, **kwargs)

# BeneficiaryCategory Views
class BeneficiaryCategoryListView(LoginRequiredMixin, AnyRoleRequiredMixin, ListView):
    model = BeneficiaryCategory
    template_name = 'categories/category_list.html'
    context_object_name = 'categories'
    paginate_by = 10

class BeneficiaryCategoryDetailView(LoginRequiredMixin, AnyRoleRequiredMixin, DetailView):
    model = BeneficiaryCategory
    template_name = 'categories/category_detail.html'
    context_object_name = 'category'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        category = self.get_object()

        # Get beneficiaries in this category
        beneficiaries = category.beneficiaries.all()

        # Calculate total annual disbursement based on beneficiaries' programs
        total_annual_disbursement = 0
        program_ids = set()

        for beneficiary in beneficiaries:
            if beneficiary.program:
                program_ids.add(beneficiary.program.id)
                # Assuming 12 months per year
                total_annual_disbursement += beneficiary.program.monthly_amount * 12

        context['total_annual_disbursement'] = total_annual_disbursement

        # Get related programs (programs that beneficiaries in this category are enrolled in)
        related_programs = Program.objects.filter(id__in=program_ids)
        context['related_programs'] = related_programs

        return context

class BeneficiaryCategoryCreateView(LoginRequiredMixin, AdminOrCaseManagerRequiredMixin, CreateView):
    model = BeneficiaryCategory
    fields = ['name', 'description', 'max_annual_amount']
    template_name = 'categories/category_form.html'
    success_url = '/categories/'

    def form_valid(self, form):
        messages.success(self.request, f"Category '{form.instance.name}' created successfully.")
        return super().form_valid(form)

class BeneficiaryCategoryUpdateView(LoginRequiredMixin, AdminOrCaseManagerRequiredMixin, UpdateView):
    model = BeneficiaryCategory
    fields = ['name', 'description', 'max_annual_amount']
    template_name = 'categories/category_form.html'
    success_url = '/categories/'

    def form_valid(self, form):
        messages.success(self.request, f"Category '{form.instance.name}' updated successfully.")
        return super().form_valid(form)

class BeneficiaryCategoryDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = BeneficiaryCategory
    template_name = 'categories/category_confirm_delete.html'
    success_url = '/categories/'

    def delete(self, request, *args, **kwargs):
        category = self.get_object()
        messages.success(request, f"Category '{category.name}' deleted successfully.")
        return super().delete(request, *args, **kwargs)

# Report Views
class ReportTemplateListView(LoginRequiredMixin, ListView):
    model = ReportTemplate
    template_name = 'reports/report_template_list.html'
    context_object_name = 'report_templates'

    def get_queryset(self):
        # Show only templates created by the user or shared templates
        return ReportTemplate.objects.filter(created_by=self.request.user)

class ReportTemplateDetailView(LoginRequiredMixin, DetailView):
    model = ReportTemplate
    template_name = 'reports/report_template_detail.html'
    context_object_name = 'report_template'

    def get_queryset(self):
        # Show only templates created by the user
        return ReportTemplate.objects.filter(created_by=self.request.user)

class ReportTemplateCreateView(LoginRequiredMixin, CreateView):
    model = ReportTemplate
    template_name = 'reports/report_template_form.html'
    fields = ['name', 'description', 'entity_type', 'fields']
    success_url = reverse_lazy('report_template_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)

class ReportTemplateUpdateView(LoginRequiredMixin, UpdateView):
    model = ReportTemplate
    template_name = 'reports/report_template_form.html'
    fields = ['name', 'description', 'entity_type', 'fields']

    def get_queryset(self):
        # Allow editing only templates created by the user
        return ReportTemplate.objects.filter(created_by=self.request.user)

    def get_success_url(self):
        return reverse('report_template_detail', kwargs={'pk': self.object.pk})

class ReportTemplateDeleteView(LoginRequiredMixin, DeleteView):
    model = ReportTemplate
    template_name = 'reports/report_template_confirm_delete.html'
    success_url = reverse_lazy('report_template_list')

    def get_queryset(self):
        # Allow deleting only templates created by the user
        return ReportTemplate.objects.filter(created_by=self.request.user)

class ReportListView(LoginRequiredMixin, ListView):
    model = Report
    template_name = 'reports/report_list.html'
    context_object_name = 'reports'

    def get_queryset(self):
        # Show only reports created by the user
        return Report.objects.filter(created_by=self.request.user)

class ReportDetailView(LoginRequiredMixin, DetailView):
    model = Report
    template_name = 'reports/report_detail.html'
    context_object_name = 'report'

    def get_queryset(self):
        # Show only reports created by the user
        return Report.objects.filter(created_by=self.request.user)

class ReportCreateView(LoginRequiredMixin, CreateView):
    model = Report
    template_name = 'reports/report_form.html'
    fields = ['name', 'template', 'filters', 'format']
    success_url = reverse_lazy('report_list')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Limit template choices to those created by the user
        form.fields['template'].queryset = ReportTemplate.objects.filter(created_by=self.request.user)
        return form

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)

class ReportUpdateView(LoginRequiredMixin, UpdateView):
    model = Report
    template_name = 'reports/report_form.html'
    fields = ['name', 'template', 'filters', 'format']

    def get_queryset(self):
        # Allow editing only reports created by the user
        return Report.objects.filter(created_by=self.request.user)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Limit template choices to those created by the user
        form.fields['template'].queryset = ReportTemplate.objects.filter(created_by=self.request.user)
        return form

    def get_success_url(self):
        return reverse('report_detail', kwargs={'pk': self.object.pk})

class ReportDeleteView(LoginRequiredMixin, DeleteView):
    model = Report
    template_name = 'reports/report_confirm_delete.html'
    success_url = reverse_lazy('report_list')

    def get_queryset(self):
        # Allow deleting only reports created by the user
        return Report.objects.filter(created_by=self.request.user)

@login_required
def generate_report(request, pk):
    """Generate a report based on the saved configuration"""
    report = get_object_or_404(Report, pk=pk, created_by=request.user)
    template = report.template

    # Get the data based on the entity type
    data = get_report_data(request, template.entity_type, report.get_filters_dict())

    # Generate the report in the requested format
    if report.format == 'excel':
        return generate_excel_report(request, report, data)
    else:
        return generate_pdf_report(request, report, data)

@login_required
def generate_custom_report(request):
    """Generate a custom report on the fly"""
    if request.method == 'POST':
        entity_type = request.POST.get('entity_type')
        fields = request.POST.getlist('fields')
        filters = json.loads(request.POST.get('filters', '{}'))
        format_type = request.POST.get('format', 'excel')

        # Create a temporary report template
        temp_template = ReportTemplate(
            name=f"Custom {entity_type.capitalize()} Report",
            entity_type=entity_type,
            fields=json.dumps(fields),
            created_by=request.user
        )

        # Get the data based on the entity type
        data = get_report_data(request, entity_type, filters)

        # Generate the report in the requested format
        if format_type == 'excel':
            return generate_excel_report(request, None, data, temp_template)
        else:
            return generate_pdf_report(request, None, data, temp_template)

    # If GET request, show the form
    entity_types = dict(ReportTemplate.ENTITY_CHOICES)

    context = {
        'entity_types': entity_types,
        'field_options': get_field_options(),
    }

    return render(request, 'reports/custom_report_form.html', context)

def get_report_data(request, entity_type, filters=None):
    """Get the data for a report based on the entity type and filters"""
    if filters is None:
        filters = {}

    # Apply role-based restrictions
    if request.user.role == 'case_manager':
        if entity_type == 'case':
            filters['case_manager'] = request.user.id
        elif entity_type == 'assessment':
            filters['created_by'] = request.user.id
        elif entity_type == 'case_note':
            filters['created_by'] = request.user.id
    elif request.user.role == 'field_officer':
        if entity_type == 'assessment':
            filters['created_by'] = request.user.id
        elif entity_type == 'case_note':
            filters['created_by'] = request.user.id

    # Get the data based on the entity type
    if entity_type == 'beneficiary':
        queryset = Beneficiary.objects.all()
    elif entity_type == 'case':
        queryset = Case.objects.all()
    elif entity_type == 'assessment':
        queryset = Assessment.objects.all()
    elif entity_type == 'case_note':
        queryset = CaseNote.objects.all()
    elif entity_type == 'program':
        queryset = Program.objects.all()
    elif entity_type == 'category':
        queryset = BeneficiaryCategory.objects.all()
    else:
        queryset = []

    # Apply filters
    for key, value in filters.items():
        if value:  # Only apply non-empty filters
            queryset = queryset.filter(**{key: value})

    return queryset

def get_field_options():
    """Get available fields for each entity type"""
    return {
        'beneficiary': [
            'id', 'name', 'dob', 'gender', 'address', 'category__name', 
            'program__name', 'created_at', 'updated_at'
        ],
        'case': [
            'id', 'title', 'beneficiary__name', 'case_manager__username', 
            'status', 'description', 'opened_date', 'closed_date', 
            'created_at', 'updated_at'
        ],
        'assessment': [
            'id', 'title', 'case__title', 'case__beneficiary__name', 
            'created_by__username', 'amount_received', 'year', 
            'created_at', 'updated_at'
        ],
        'case_note': [
            'id', 'case__title', 'case__beneficiary__name', 
            'created_by__username', 'content', 'created_at', 'updated_at'
        ],
        'program': [
            'id', 'name', 'description', 'monthly_amount', 
            'next_program__name'
        ],
        'category': [
            'id', 'name', 'description', 'max_annual_amount'
        ]
    }

def generate_excel_report(request, report=None, data=None, template=None):
    """Generate an Excel report"""
    if report:
        template = report.template

    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = template.name[:31]  # Excel worksheet names are limited to 31 chars

    # Get the fields to include
    fields = template.get_fields_list()

    # Add header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=field.replace('__', ' ').replace('_', ' ').title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add data rows
    for row_idx, item in enumerate(data, 2):
        for col_idx, field in enumerate(fields, 1):
            # Handle nested fields (e.g., beneficiary__name)
            if '__' in field:
                parts = field.split('__')
                value = item
                for part in parts:
                    if hasattr(value, part):
                        value = getattr(value, part)
                    else:
                        value = None
                        break
            else:
                value = getattr(item, field, None)

            ws.cell(row=row_idx, column=col_idx, value=value)

    # Add summary row
    summary_row = len(data) + 3
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=f"Total Records: {len(data)}")

    # Auto-adjust column widths
    for col_idx, _ in enumerate(fields, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{template.name}.xlsx"'

    # Save the workbook to the response
    wb.save(response)
    return response

def generate_pdf_report(request, report=None, data=None, template=None):
    """Generate a PDF report"""
    if report:
        template = report.template

    # Get the fields to include
    fields = template.get_fields_list()

    # Prepare data for the template
    headers = [field.replace('__', ' ').replace('_', ' ').title() for field in fields]
    rows = []

    for item in data:
        row = []
        for field in fields:
            # Handle nested fields (e.g., beneficiary__name)
            if '__' in field:
                parts = field.split('__')
                value = item
                for part in parts:
                    if hasattr(value, part):
                        value = getattr(value, part)
                    else:
                        value = None
                        break
            else:
                value = getattr(item, field, None)

            row.append(value)
        rows.append(row)

    # Prepare context for the template
    context = {
        'title': template.name,
        'description': template.description,
        'headers': headers,
        'rows': rows,
        'total_records': len(data),
        'generated_by': request.user.username,
        'generated_at': timezone.now(),
    }

    # Render the template
    html_string = get_template('reports/pdf_report_template.html').render(context)

    # Create PDF
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{template.name}.pdf"'
        return response

    return HttpResponse('Error generating PDF', status=400)


# Action Plan Views
class ActionPlanListView(LoginRequiredMixin, AdminOrCaseManagerRequiredMixin, ListView):
    model = ActionPlan
    template_name = 'action_plans/action_plan_list.html'
    context_object_name = 'action_plans'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filter action plans based on user role
        if self.request.user.role == 'case_manager':
            # Case managers can see action plans they created
            queryset = queryset.filter(created_by=self.request.user)

        # Apply search filter
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(title__icontains=search_query) | 
                Q(description__icontains=search_query) |
                Q(goals__icontains=search_query) |
                Q(case__title__icontains=search_query) |
                Q(case__beneficiary__name__icontains=search_query)
            )

        # Apply status filter
        status = self.request.GET.get('status', '')
        if status:
            queryset = queryset.filter(status=status)

        # Apply sorting
        sort = self.request.GET.get('sort', '-created_at')
        queryset = queryset.order_by(sort)

        return queryset


class ActionPlanDetailView(LoginRequiredMixin, AdminOrCaseManagerRequiredMixin, DetailView):
    model = ActionPlan
    template_name = 'action_plans/action_plan_detail.html'
    context_object_name = 'action_plan'

    def get_queryset(self):
        queryset = super().get_queryset()
        # Case managers can only view their own action plans
        if self.request.user.role == 'case_manager':
            queryset = queryset.filter(created_by=self.request.user)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        action_plan = self.get_object()

        # Get related progress records
        context['progress_records'] = BeneficiaryProgress.objects.filter(
            action_plan=action_plan
        ).order_by('-date')

        return context


class ActionPlanCreateView(LoginRequiredMixin, AdminOrCaseManagerRequiredMixin, CreateView):
    model = ActionPlan
    template_name = 'action_plans/action_plan_form.html'
    fields = ['title', 'case', 'description', 'goals', 'timeline', 'status']

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Limit case choices based on user role
        if self.request.user.role == 'case_manager':
            form.fields['case'].queryset = Case.objects.filter(case_manager=self.request.user)
        return form

    def form_valid(self, form):
        # Set the creator to the current user
        form.instance.created_by = self.request.user
        messages.success(self.request, f"Action plan '{form.instance.title}' created successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('action_plan_detail', kwargs={'pk': self.object.pk})


class ActionPlanUpdateView(LoginRequiredMixin, AdminOrCaseManagerRequiredMixin, UpdateView):
    model = ActionPlan
    template_name = 'action_plans/action_plan_form.html'
    fields = ['title', 'case', 'description', 'goals', 'timeline', 'status']

    def get_queryset(self):
        queryset = super().get_queryset()
        # Case managers can only update their own action plans
        if self.request.user.role == 'case_manager':
            queryset = queryset.filter(created_by=self.request.user)
        return queryset

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Limit case choices based on user role
        if self.request.user.role == 'case_manager':
            form.fields['case'].queryset = Case.objects.filter(case_manager=self.request.user)
        return form

    def form_valid(self, form):
        messages.success(self.request, f"Action plan '{form.instance.title}' updated successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('action_plan_detail', kwargs={'pk': self.object.pk})


class ActionPlanDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = ActionPlan
    template_name = 'action_plans/action_plan_confirm_delete.html'
    context_object_name = 'action_plan'
    success_url = reverse_lazy('action_plan_list')

    def delete(self, request, *args, **kwargs):
        action_plan = self.get_object()
        messages.success(request, f"Action plan '{action_plan.title}' deleted successfully.")
        return super().delete(request, *args, **kwargs)


# Beneficiary Progress Views
class BeneficiaryProgressListView(LoginRequiredMixin, CanTrackBeneficiaryProgressMixin, ListView):
    model = BeneficiaryProgress
    template_name = 'progress/progress_list.html'
    context_object_name = 'progress_records'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filter progress records based on user role
        if self.request.user.role == 'case_manager':
            # Case managers can see progress for their cases
            case_ids = Case.objects.filter(case_manager=self.request.user).values_list('id', flat=True)
            queryset = queryset.filter(case__id__in=case_ids)

        # Apply beneficiary filter
        beneficiary_id = self.request.GET.get('beneficiary', '')
        if beneficiary_id:
            queryset = queryset.filter(beneficiary_id=beneficiary_id)

        # Apply case filter
        case_id = self.request.GET.get('case', '')
        if case_id:
            queryset = queryset.filter(case_id=case_id)

        # Apply action plan filter
        action_plan_id = self.request.GET.get('action_plan', '')
        if action_plan_id:
            queryset = queryset.filter(action_plan_id=action_plan_id)

        # Apply status filter
        status = self.request.GET.get('status', '')
        if status:
            queryset = queryset.filter(status=status)

        # Apply search filter
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(beneficiary__name__icontains=search_query) | 
                Q(notes__icontains=search_query) |
                Q(case__title__icontains=search_query) |
                Q(action_plan__title__icontains=search_query)
            )

        # Apply sorting
        sort = self.request.GET.get('sort', '-date')
        queryset = queryset.order_by(sort)

        return queryset


class BeneficiaryProgressDetailView(LoginRequiredMixin, CanTrackBeneficiaryProgressMixin, DetailView):
    model = BeneficiaryProgress
    template_name = 'progress/progress_detail.html'
    context_object_name = 'progress'

    def get_queryset(self):
        queryset = super().get_queryset()
        # Filter based on user role
        if self.request.user.role == 'case_manager':
            # Case managers can see progress for their cases
            case_ids = Case.objects.filter(case_manager=self.request.user).values_list('id', flat=True)
            queryset = queryset.filter(case__id__in=case_ids)
        return queryset


class BeneficiaryProgressCreateView(LoginRequiredMixin, CanTrackBeneficiaryProgressMixin, CreateView):
    model = BeneficiaryProgress
    template_name = 'progress/progress_form.html'
    fields = ['beneficiary', 'case', 'action_plan', 'date', 'status', 'progress_percentage', 'notes']

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Limit choices based on user role
        if self.request.user.role == 'case_manager':
            # Case managers can only select their cases
            case_ids = Case.objects.filter(case_manager=self.request.user).values_list('id', flat=True)
            form.fields['case'].queryset = Case.objects.filter(id__in=case_ids)

            # Limit beneficiaries to those in their cases
            beneficiary_ids = Case.objects.filter(case_manager=self.request.user).values_list('beneficiary_id', flat=True)
            form.fields['beneficiary'].queryset = Beneficiary.objects.filter(id__in=beneficiary_ids)

            # Limit action plans to those they created
            form.fields['action_plan'].queryset = ActionPlan.objects.filter(created_by=self.request.user)
        return form

    def form_valid(self, form):
        # Set the recorder to the current user
        form.instance.recorded_by = self.request.user
        messages.success(self.request, f"Progress record for {form.instance.beneficiary.name} created successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('beneficiary_progress_detail', kwargs={'pk': self.object.pk})


class BeneficiaryProgressUpdateView(LoginRequiredMixin, CanTrackBeneficiaryProgressMixin, UpdateView):
    model = BeneficiaryProgress
    template_name = 'progress/progress_form.html'
    fields = ['beneficiary', 'case', 'action_plan', 'date', 'status', 'progress_percentage', 'notes']

    def get_queryset(self):
        queryset = super().get_queryset()
        # Filter based on user role
        if self.request.user.role == 'case_manager':
            # Case managers can only update progress for their cases
            case_ids = Case.objects.filter(case_manager=self.request.user).values_list('id', flat=True)
            queryset = queryset.filter(case__id__in=case_ids)
        return queryset

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Limit choices based on user role
        if self.request.user.role == 'case_manager':
            # Case managers can only select their cases
            case_ids = Case.objects.filter(case_manager=self.request.user).values_list('id', flat=True)
            form.fields['case'].queryset = Case.objects.filter(id__in=case_ids)

            # Limit beneficiaries to those in their cases
            beneficiary_ids = Case.objects.filter(case_manager=self.request.user).values_list('beneficiary_id', flat=True)
            form.fields['beneficiary'].queryset = Beneficiary.objects.filter(id__in=beneficiary_ids)

            # Limit action plans to those they created
            form.fields['action_plan'].queryset = ActionPlan.objects.filter(created_by=self.request.user)
        return form

    def form_valid(self, form):
        messages.success(self.request, f"Progress record for {form.instance.beneficiary.name} updated successfully.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('beneficiary_progress_detail', kwargs={'pk': self.object.pk})


class BeneficiaryProgressDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = BeneficiaryProgress
    template_name = 'progress/progress_confirm_delete.html'
    context_object_name = 'progress'
    success_url = reverse_lazy('beneficiary_progress_list')

    def delete(self, request, *args, **kwargs):
        progress = self.get_object()
        messages.success(request, f"Progress record for {progress.beneficiary.name} deleted successfully.")
        return super().delete(request, *args, **kwargs)
